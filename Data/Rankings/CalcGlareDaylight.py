# =========== Lux => W =============

# 1 Sun (1,000 W/m2) equals approximately 120,000 Lux
# Src: https://ieee-dataport.org/open-access/conversion-guide-solar-irradiance-and-lux-illuminance

# Solar Irradiance of 1 Sun (1000 W/m2) ... (122 ± 1) klx for outdoor sunlight
# Src: https://www.jvejournals.com/article/21667

# =========== Daylight Range =============

# Daylighting between 100 and 3000 lux
# Src: https://www.ies.org/definitions/useful-daylight-illuminance-udi/

# Good range to consider: 100 to 5000 lux
# Src: https://www.engineeringtoolbox.com/light-level-rooms-d_708.html

# You ever get 3/4's thru something and realize you should've done it in R?
# also openpyxl was dumb, shouldve just used csv...
# whatever, it's done anyway

import sys
from openpyxl import load_workbook
from openpyxl import worksheet
import csv # for output, writing to EXCEL is a pain, just CSV much easier

csvs = ["conf_Pos1.xlsx", "office_Pos1.xlsx"]
daylighting_range = (100.0, 5000.0) # 100 and 3,000 lux

def main():
	for csv_name in csvs:
		out_facades = []
		out_total = []
		out_daylight = []
		out_glare = []
		try:
			wb = load_workbook(csv_name)
			for sheet_name in wb:
				if("Sheet" in sheet_name.title): # Skip default sheet
					print("Skipping sheet \"" + sheet_name.title + "\"")
					continue
				ws = wb[sheet_name.title]
				num_cells = ws.max_row # Takes into account header and index starts at 1 so it cancels out
				daylit_cells = 0
				glare_cells = 0
				for i in range(2, num_cells + 1): # skip header, also index starts at 1 instead of zero for some reason in this function
					val_kWh = ws.cell(row=i, column=1, value=None).value
					val_lux = kWh_to_lux(val_kWh)
					if(daylighting_range[0] <= val_lux <= daylighting_range[1]):
						daylit_cells = daylit_cells + 1
					elif(val_lux > daylighting_range[1]):
						glare_cells = glare_cells + 1
				print(sheet_name.title)
				print("Total:\n" + str(num_cells))
				print("Daylit:\n" + str(daylit_cells))
				print("Glare:\n" + str(glare_cells))
				out_facades.append(sheet_name.title)
				out_total.append(num_cells)
				out_daylight.append(daylit_cells)
				out_glare.append(glare_cells)
		except Exception as e:
			raise e
			sys.exit()
		# Write info gleaned to CSV for easy copy paste to Google Sheet
		with open("CalcDaylightOutput\\out_" + csv_name[:-5] + ".csv", 'w', newline='') as csvfile:
			writer = csv.writer(csvfile)
			writer.writerow(["Facade", "Glare", "Total", "Daylit", "Total"])
			for i in range(0, len(out_facades)):
				writer.writerow([out_facades[i], out_glare[i], out_total[i], out_daylight[i], out_total[i]])

def kWh_to_lux(kWh):
	return kWh * 122000

if __name__ == "__main__":
	main()